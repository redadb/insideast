#!/usr/bin/env python3
"""
Create styled Excel with category organization
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the existing Excel file
df = pd.read_excel('/Users/redadrissielbouzaidi/Downloads/insideast/Trust_Elements.xlsx')

# Categorize entries more accurately
def categorize(row):
    category = str(row['Category']).lower() if pd.notna(row['Category']) else ''
    name = str(row['Name/Handle']).lower()
    
    if 'interior design' in category or 'interior design' in name:
        return 'Interior Designers'
    elif 'hotel' in category or 'hotel' in name or 'resort' in name:
        return 'Hotels & Resorts'
    elif 'airbnb' in category or 'airbnb' in name or 'vacation' in name:
        return 'Vacation Rentals'
    elif 'architect' in category or 'architect' in name:
        return 'Architects'
    elif 'home decor' in category or 'home improvement' in category:
        return 'Home Decor & DIY'
    elif 'artist' in category or 'photographer' in category:
        return 'Artists & Photographers'
    elif 'digital creator' in category or 'personal blog' in category or 'influencer' in category:
        return 'Content Creators & Influencers'
    elif 'plumbing' in category or 'renovations' in category or 'custom builds' in category:
        return 'Construction & Renovation'
    else:
        return 'Other Customers'

df['Category_Group'] = df.apply(categorize, axis=1)

# Sort by category group and then by followers
def extract_follower_number(val):
    import re
    if pd.isna(val) or val == '':
        return 0
    match = re.search(r'(\d+\.?\d*)', str(val))
    if match:
        num = float(match.group(1))
        if 'k' in str(val).lower():
            return num * 1000
        return num
    return 0

df['follower_sort'] = df['Followers'].apply(extract_follower_number)
df = df.sort_values(['Category_Group', 'follower_sort'], ascending=[True, False])

# Reorder columns
columns_order = ['Category_Group', 'Name/Handle', 'Followers', 'Category', 'Location', 'Website', 'Total Posts', 'Instagram Posts', 'Testimonial/Notes']
df = df[columns_order]

# Create Excel with styling
output_file = '/Users/redadrissielbouzaidi/Downloads/insideast/Trust_Elements_Styled.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Influencers & Customers')

# Load workbook and apply styling
wb = load_workbook(output_file)
ws = wb['Influencers & Customers']

# Define styles
header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

category_fills = {
    'Interior Designers': PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid'),
    'Hotels & Resorts': PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid'),
    'Vacation Rentals': PatternFill(start_color='F0F8E8', end_color='F0F8E8', fill_type='solid'),
    'Architects': PatternFill(start_color='F8E8F4', end_color='F8E8F4', fill_type='solid'),
    'Home Decor & DIY': PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid'),
    'Artists & Photographers': PatternFill(start_color='F4F0FF', end_color='F4F0FF', fill_type='solid'),
    'Content Creators & Influencers': PatternFill(start_color='FFF8E8', end_color='FFF8E8', fill_type='solid'),
    'Construction & Renovation': PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type='solid'),
    'Other Customers': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
}

category_header_font = Font(name='Calibri', size=10, bold=True, color='1F4788')

border_thin = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Style header row
for col_num, cell in enumerate(ws[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = alignment_center
    cell.border = border_thin

# Set column widths
ws.column_dimensions['A'].width = 25  # Category Group
ws.column_dimensions['B'].width = 35  # Name
ws.column_dimensions['C'].width = 12  # Followers
ws.column_dimensions['D'].width = 28  # Category
ws.column_dimensions['E'].width = 18  # Location
ws.column_dimensions['F'].width = 38  # Website
ws.column_dimensions['G'].width = 10  # Total Posts
ws.column_dimensions['H'].width = 45  # Instagram Posts
ws.column_dimensions['I'].width = 50  # Testimonial

# Apply styling to data rows
current_category = None
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
    category_val = row[0].value  # Category_Group column
    
    # Get appropriate fill color
    fill = category_fills.get(category_val, PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'))
    
    for col_num, cell in enumerate(row, 1):
        cell.border = border_thin
        cell.fill = fill
        
        # Column A (Category Group) - Bold if new category
        if col_num == 1:
            if category_val != current_category:
                cell.font = category_header_font
                current_category = category_val
            cell.alignment = alignment_left
        
        # Column B (Name) - Left align
        elif col_num == 2:
            cell.alignment = alignment_left
            cell.font = Font(name='Calibri', size=10)
        
        # Column C (Followers) - Center align, bold
        elif col_num == 3:
            cell.alignment = alignment_center
            if cell.value:
                cell.font = Font(name='Calibri', size=10, bold=True, color='D35400')
        
        # Column G (Total Posts) - Center align
        elif col_num == 7:
            cell.alignment = alignment_center
        
        # All other columns - Left align
        else:
            cell.alignment = alignment_left
            cell.font = Font(name='Calibri', size=9)

# Set row heights
ws.row_dimensions[1].height = 25  # Header row
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 30

# Freeze first row and first column
ws.freeze_panes = 'B2'

# Save
wb.save(output_file)

print(f"✅ Styled Excel created: {output_file}")
print(f"\n📊 Category breakdown:")
category_counts = df['Category_Group'].value_counts()
for cat, count in category_counts.items():
    print(f"  • {cat}: {count} entries")
